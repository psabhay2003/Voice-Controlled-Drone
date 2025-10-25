import pandas as pd
from openpyxl import load_workbook

excel_file = "VCD_DATA.xlsx"  
sheet_name = "Sheet1"        

wb = load_workbook(excel_file, data_only=True)
ws = wb[sheet_name]

data = []

for row in ws.iter_rows(min_row=2):  
    name = row[2].value              
    cell = row[5]                    
    # Extract hyperlink (if exists)
    if cell.hyperlink:
        link = cell.hyperlink.target
    else:
        link = cell.value            

    data.append((name, link))

df = pd.DataFrame(data, columns=["audio_name", "drive_link"])
df.to_csv("dataset_links.csv", index=False)

print("✅ Extracted hyperlinks to dataset_links.csv")
